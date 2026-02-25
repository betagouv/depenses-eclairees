"""
Extraction de texte depuis les fichiers Excel : xlsx, xls, ods.
Retourne un contenu markdown (tableaux avec |, légende des fusions, ## par feuille).
"""

import io

import pandas as pd
import xlrd
from openpyxl import load_workbook

MERGE_LEGEND = "Légende : **#** = cellule appartenant à une plage fusionnée verticalement (la valeur figure dans la première cellule en haut de la plage)."


def _merged_continuation_cells(ws) -> set[tuple[int, int]]:
    """Cellules fusionnées verticalement (hors cellule d'origine). Coordonnées 1-based (openpyxl)."""
    continuation: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        min_row, max_row = merged_range.min_row, merged_range.max_row
        if max_row <= min_row:
            continue
        min_col = merged_range.min_col
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (min_row, min_col):
                    continuation.add((row, col))
    return continuation


def _sheet_to_dataframe_with_merge_markers(ws) -> pd.DataFrame:
    """Lit une feuille openpyxl. Cellules fusionnées verticalement : valeur en haut, '#' ailleurs."""
    continuation = _merged_continuation_cells(ws)
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return pd.DataFrame()

    rows = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            if (r, c) in continuation:
                row_vals.append("#")
            else:
                cell = ws.cell(row=r, column=c)
                row_vals.append(cell.value)
        rows.append(row_vals)
    return pd.DataFrame(rows)


def _drop_empty_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Supprime les lignes dont toutes les cellules sont vides."""
    if df.empty:
        return df
    mask = df.apply(
        lambda row: any(not pd.isna(v) and str(v).strip() != "" for v in row),
        axis=1,
    )
    return df.loc[mask].reset_index(drop=True)


def _drop_trailing_empty_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Supprime les colonnes entièrement vides à droite."""
    if df.empty:
        return df
    for col_idx in range(len(df.columns) - 1, -1, -1):
        col = df.iloc[:, col_idx]
        if any(not pd.isna(v) and str(v).strip() != "" for v in col):
            break
    else:
        return pd.DataFrame()
    return df.iloc[:, : col_idx + 1]


def _dataframe_to_markdown_pipe(df: pd.DataFrame) -> str:
    """Convertit un DataFrame en table markdown avec |."""
    if df.empty:
        return ""
    df = df.fillna("")
    lines = []
    for _, row in df.iterrows():
        cells = [str(v).strip().replace("\r\n", " ").replace("\n", " ").replace("\r", " ") for v in row]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _xlsx_sheet_to_markdown(ws) -> str:
    """Contenu d'une feuille openpyxl en markdown."""
    df = _sheet_to_dataframe_with_merge_markers(ws)
    if df.empty:
        return ""
    df = _drop_empty_rows(df)
    if df.empty:
        return ""
    df = _drop_trailing_empty_columns(df)
    if df.empty:
        return ""
    return _dataframe_to_markdown_pipe(df)


def _ods_sheet_to_markdown(df: pd.DataFrame) -> str:
    """Feuille ODS (DataFrame) en markdown."""
    if df.empty:
        return ""
    df = _drop_empty_rows(df)
    if df.empty:
        return ""
    df = _drop_trailing_empty_columns(df)
    if df.empty:
        return ""
    return _dataframe_to_markdown_pipe(df)


def _xls_merged_continuation_cells(sheet) -> set[tuple[int, int]]:
    """Cellules fusionnées verticalement (xlrd). Coordonnées 0-based."""
    continuation: set[tuple[int, int]] = set()
    if not hasattr(sheet, "merged_cells"):
        return continuation
    for (rlo, rhi, clo, chi) in sheet.merged_cells:
        if rhi - rlo <= 1:
            continue
        for r in range(rlo + 1, rhi):
            for c in range(clo, chi):
                continuation.add((r, c))
    return continuation


def _xls_sheet_to_dataframe(sheet) -> pd.DataFrame:
    """Feuille xlrd → DataFrame avec marqueurs de fusion."""
    continuation = _xls_merged_continuation_cells(sheet)
    nrows = sheet.nrows
    ncols = sheet.ncols
    if nrows == 0 or ncols == 0:
        return pd.DataFrame()

    rows = []
    for r in range(nrows):
        row_vals = []
        for c in range(ncols):
            if (r, c) in continuation:
                row_vals.append("#")
            else:
                row_vals.append(sheet.cell_value(r, c))
        rows.append(row_vals)
    return pd.DataFrame(rows)


def _xls_sheet_to_markdown(sheet) -> str:
    """Feuille xlrd en markdown."""
    df = _xls_sheet_to_dataframe(sheet)
    if df.empty:
        return ""
    df = _drop_empty_rows(df)
    if df.empty:
        return ""
    df = _drop_trailing_empty_columns(df)
    if df.empty:
        return ""
    return _dataframe_to_markdown_pipe(df)


def extract_text_from_xlsx(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier XLSX à partir de son contenu binaire.

    Args:
        file_content: Contenu du fichier .xlsx
        file_path: Chemin ou nom du fichier (pour les logs)
        sep: Séparateur entre les feuilles

    Returns:
        (texte markdown, False) — pas d'OCR pour Excel
    """
    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=False, data_only=True)
    except Exception as e:
        print(f"Erreur lecture XLSX {file_path}: {e}")
        return "", False

    try:
        sheet_names = wb.sheetnames
        parts = [MERGE_LEGEND]
        for name in sheet_names:
            ws = wb[name]
            md = _xlsx_sheet_to_markdown(ws)
            if md:
                parts.append(f"## {name}\n\n{md}")
        wb.close()
        return sep.join(parts), False
    except Exception as e:
        print(f"Erreur extraction XLSX {file_path}: {e}")
        return "", False


def extract_text_from_ods(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier ODS à partir de son contenu binaire.

    Returns:
        (texte markdown, False)
    """
    try:
        all_sheets: dict = pd.read_excel(io.BytesIO(file_content), engine="odf", sheet_name=None)
    except Exception as e:
        print(f"Erreur lecture ODS {file_path}: {e}")
        return "", False

    try:
        parts = [MERGE_LEGEND]
        for name, df in all_sheets.items():
            md = _ods_sheet_to_markdown(df)
            if md:
                parts.append(f"## {name}\n\n{md}")
        return sep.join(parts), False
    except Exception as e:
        print(f"Erreur extraction ODS {file_path}: {e}")
        return "", False


def extract_text_from_xls(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier XLS à partir de son contenu binaire.

    Returns:
        (texte markdown, False)
    """
    try:
        wb = xlrd.open_workbook(file_contents=file_content, formatting_info=True)
    except (xlrd.XLRDError, NotImplementedError):
        try:
            wb = xlrd.open_workbook(file_contents=file_content, formatting_info=False)
        except Exception as e:
            print(f"Erreur lecture XLS {file_path}: {e}")
            return "", False

    try:
        sheet_names = wb.sheet_names()
        parts = [MERGE_LEGEND]
        for name in sheet_names:
            try:
                sheet = wb.sheet_by_name(name)
            except xlrd.XLRDError:
                continue
            md = _xls_sheet_to_markdown(sheet)
            if md:
                parts.append(f"## {name}\n\n{md}")
        return sep.join(parts), False
    except Exception as e:
        print(f"Erreur extraction XLS {file_path}: {e}")
        return "", False
