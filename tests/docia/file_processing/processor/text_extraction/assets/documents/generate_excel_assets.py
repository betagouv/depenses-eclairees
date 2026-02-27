#!/usr/bin/env python3
"""
Génère les fichiers de test Excel (XLSX, XLS, ODS) avec plusieurs onglets et cellules fusionnées.
À exécuter depuis la racine du projet ou depuis ce répertoire :
  python tests/docia/file_processing/processor/text_extraction/assets/documents/generate_excel_assets.py
Les fichiers sample.xlsx, sample.xls, sample.ods sont créés dans ce répertoire.
ODS : généré sans odfpy (stdlib : zipfile + XML).
"""

import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

OUT_DIR = Path(__file__).resolve().parent

# Namespaces ODF (OASIS) pour construire content.xml
_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}


def _odf_tag(ns: str, local: str) -> str:
    return "{%s}%s" % (_NS[ns], local)


def _ods_build_content_xml() -> bytes:
    """Construit le XML content.xml d'un ODS (2 feuilles, cellules fusionnées)."""
    root = ET.Element(
        _odf_tag("office", "document-content"),
        attrib={
            _odf_tag("office", "mimetype"): "application/vnd.oasis.opendocument.spreadsheet",
            _odf_tag("office", "version"): "1.2",
        },
    )
    body = ET.SubElement(root, _odf_tag("office", "body"))
    spreadsheet = ET.SubElement(body, _odf_tag("office", "spreadsheet"))

    def cell_elt(parent, text: str, cols: int = 1, rows: int = 1):
        attrs = {_odf_tag("office", "value-type"): "string"}
        if cols > 1:
            attrs[_odf_tag("table", "number-columns-spanned")] = str(cols)
        if rows > 1:
            attrs[_odf_tag("table", "number-rows-spanned")] = str(rows)
        elt = ET.SubElement(parent, _odf_tag("table", "table-cell"), attrib=attrs)
        p = ET.SubElement(elt, _odf_tag("text", "p"))
        p.text = text
        return elt

    def covered_cell(parent):
        return ET.SubElement(parent, _odf_tag("table", "covered-table-cell"))

    # Feuille 1
    t1 = ET.SubElement(spreadsheet, _odf_tag("table", "table"), attrib={_odf_tag("table", "name"): "Feuille1"})
    tr = ET.SubElement(t1, _odf_tag("table", "table-row"))
    cell_elt(tr, "Titre fusionné", cols=2)
    tr = ET.SubElement(t1, _odf_tag("table", "table-row"))
    cell_elt(tr, "Col A")
    cell_elt(tr, "Col B")
    for label, val in [("Ligne 1", "10"), ("Ligne 2", "20")]:
        tr = ET.SubElement(t1, _odf_tag("table", "table-row"))
        cell_elt(tr, label)
        cell_elt(tr, val)
    tr = ET.SubElement(t1, _odf_tag("table", "table-row"))
    cell_elt(tr, "Total", cols=2)

    # Feuille 2 (fusion verticale)
    t2 = ET.SubElement(spreadsheet, _odf_tag("table", "table"), attrib={_odf_tag("table", "name"): "Feuille2"})
    tr = ET.SubElement(t2, _odf_tag("table", "table-row"))
    cell_elt(tr, "Section", rows=3)
    cell_elt(tr, "V1")
    tr = ET.SubElement(t2, _odf_tag("table", "table-row"))
    covered_cell(tr)
    cell_elt(tr, "V2")
    tr = ET.SubElement(t2, _odf_tag("table", "table-row"))
    covered_cell(tr)
    cell_elt(tr, "V3")
    tr = ET.SubElement(t2, _odf_tag("table", "table-row"))
    cell_elt(tr, "Fin")

    return ET.tostring(root, encoding="unicode", method="xml").encode("utf-8")


def write_xlsx():
    """Crée sample.xlsx avec 2 onglets et cellules fusionnées (openpyxl)."""
    from openpyxl import Workbook

    wb = Workbook()
    # Feuille 1
    ws1 = wb.active
    ws1.title = "Feuille1"
    ws1["A1"] = "Titre fusionné"
    ws1.merge_cells("A1:B1")
    ws1["A2"], ws1["B2"] = "Col A", "Col B"
    ws1["A3"], ws1["B3"] = "Ligne 1", "10"
    ws1["A4"], ws1["B4"] = "Ligne 2", "20"
    ws1.merge_cells("A5:B5")  # fusion horizontale
    ws1["A5"] = "Total"
    # Feuille 2
    ws2 = wb.create_sheet("Feuille2")
    ws2["A1"] = "Section"
    ws2.merge_cells("A1:A3")  # fusion verticale -> doit donner # en A2, A3
    ws2["B1"], ws2["B2"], ws2["B3"] = "V1", "V2", "V3"
    ws2["A4"] = "Fin"
    wb.save(OUT_DIR / "sample.xlsx")
    print("Créé:", OUT_DIR / "sample.xlsx")


def write_xls():
    """Crée sample.xls avec 2 onglets et cellules fusionnées (xlwt)."""
    try:
        import xlwt
    except ImportError:
        print("xlwt non installé (pip install xlwt). sample.xls non généré.")
        return
    wb = xlwt.Workbook()
    # Feuille 1
    ws1 = wb.add_sheet("Feuille1")
    ws1.write_merge(0, 0, 0, 1, "Titre fusionné")
    ws1.write(1, 0, "Col A")
    ws1.write(1, 1, "Col B")
    ws1.write(2, 0, "Ligne 1")
    ws1.write(2, 1, 10)
    ws1.write(3, 0, "Ligne 2")
    ws1.write(3, 1, 20)
    ws1.write_merge(4, 4, 0, 1, "Total")
    # Feuille 2
    ws2 = wb.add_sheet("Feuille2")
    ws2.write_merge(0, 2, 0, 0, "Section")  # fusion verticale (ne pas écrire dans les cellules fusionnées)
    ws2.write(0, 1, "V1")
    ws2.write(1, 1, "V2")
    ws2.write(2, 1, "V3")
    ws2.write(3, 0, "Fin")
    wb.save(str(OUT_DIR / "sample.xls"))
    print("Créé:", OUT_DIR / "sample.xls")


def _ods_manifest_xml() -> bytes:
    """Manifest minimal pour un ODS (META-INF/manifest.xml)."""
    return b"""<?xml version="1.0" encoding="UTF-8"?>
<manifest:manifest xmlns:manifest="urn:oasis:names:tc:opendocument:xmlns:manifest:1.0">
  <manifest:file-entry manifest:media-type="application/vnd.oasis.opendocument.spreadsheet" manifest:full-path="/"/>
  <manifest:file-entry manifest:media-type="text/xml" manifest:full-path="content.xml"/>
</manifest:manifest>"""


def write_ods():
    """Crée sample.ods avec 2 onglets et cellules fusionnées (stdlib : zipfile + XML)."""
    out_path = OUT_DIR / "sample.ods"
    mimetype = b"application/vnd.oasis.opendocument.spreadsheet"
    content_xml = _ods_build_content_xml()
    manifest = _ods_manifest_xml()

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        # mimetype en premier, non compressé (recommandation ODF)
        zf.writestr("mimetype", mimetype, compress_type=zipfile.ZIP_STORED)
        zf.writestr("content.xml", content_xml)
        zf.writestr("META-INF/manifest.xml", manifest)
    print("Créé:", out_path)


if __name__ == "__main__":
    write_xlsx()
    write_xls()
    write_ods()
