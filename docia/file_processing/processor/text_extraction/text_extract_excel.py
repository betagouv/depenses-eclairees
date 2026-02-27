"""
Extraction de texte depuis les fichiers Excel : xlsx, xls, ods.
Retourne un contenu markdown (tableaux avec |, légende des fusions, ## par feuille).
Sans pandas : listes Python + openpyxl (xlsx), xlrd (xls). ODS : stdlib uniquement (zip + XML).
"""

import io
import xml.etree.ElementTree as ET
import zipfile

import xlrd
from openpyxl import load_workbook

MERGE_LEGEND = """Légende : **#** = cellule appartenant à une plage fusionnée verticalement 
(la valeur figure dans la première cellule en haut de la plage)."""

# Type pour une feuille : liste de lignes, chaque ligne = liste de valeurs de cellules
SheetRows = list[list]


def _is_empty(v) -> bool:
    """Valeur considérée comme vide (sans pandas)."""
    if v is None:
        return True
    if isinstance(v, float):
        return v != v  # NaN
    return str(v).strip() == ""


def _merged_continuation_cells(ws) -> set[tuple[int, int]]:
    """Cellules fusionnées verticalement (hors cellule d'origine). Coordonnées 1-based (openpyxl)."""
    continuation: set[tuple[int, int]] = set()
    for merged_range in ws.merged_cells.ranges:
        min_row, max_row = merged_range.min_row, merged_range.max_row
        if max_row <= min_row:
            continue
        min_col = merged_range.min_col
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                if (row, col) != (min_row, min_col):
                    continuation.add((row, col))
    return continuation


def _sheet_to_rows_with_merge_markers(ws) -> SheetRows:
    """Lit une feuille openpyxl. Cellules fusionnées verticalement : valeur en haut, '#' ailleurs."""
    continuation = _merged_continuation_cells(ws)
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    if max_row == 0 or max_col == 0:
        return []

    rows: SheetRows = []
    for r in range(1, max_row + 1):
        row_vals = []
        for c in range(1, max_col + 1):
            if (r, c) in continuation:
                row_vals.append("#")
            else:
                cell = ws.cell(row=r, column=c)
                row_vals.append(cell.value)
        rows.append(row_vals)
    return rows


def _drop_empty_rows(rows: SheetRows) -> SheetRows:
    """Supprime les lignes dont toutes les cellules sont vides."""
    if not rows:
        return []
    return [row for row in rows if any(not _is_empty(v) for v in row)]


def _drop_trailing_empty_columns(rows: SheetRows) -> SheetRows:
    """Supprime les colonnes entièrement vides à droite."""
    if not rows:
        return []
    ncols = max(len(r) for r in rows) if rows else 0
    if ncols == 0:
        return []
    for col_idx in range(ncols - 1, -1, -1):
        if any((not _is_empty(row[col_idx]) if col_idx < len(row) else False) for row in rows):
            return [r[: col_idx + 1] for r in rows]
    return []


def _rows_to_markdown_pipe(rows: SheetRows) -> str:
    """Convertit des lignes de cellules en table markdown avec |."""
    if not rows:
        return ""
    ncols = max(len(r) for r in rows)
    lines = []
    for row in rows:
        # Pad à ncols pour tableaux rectangulaires
        padded = (row + [""] * ncols)[:ncols]
        cells = [
            ("" if _is_empty(v) else str(v).strip()).replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
            for v in padded
        ]
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)


def _xlsx_sheet_to_markdown(ws) -> str:
    """Contenu d'une feuille openpyxl en markdown."""
    rows = _sheet_to_rows_with_merge_markers(ws)
    if not rows:
        return ""
    rows = _drop_empty_rows(rows)
    if not rows:
        return ""
    rows = _drop_trailing_empty_columns(rows)
    if not rows:
        return ""
    return _rows_to_markdown_pipe(rows)


def _rows_to_markdown(rows: SheetRows) -> str:
    """Feuille (liste de lignes) en markdown."""
    if not rows:
        return ""
    rows = _drop_empty_rows(rows)
    if not rows:
        return ""
    rows = _drop_trailing_empty_columns(rows)
    if not rows:
        return ""
    return _rows_to_markdown_pipe(rows)


def _xls_merged_continuation_cells(sheet) -> set[tuple[int, int]]:
    """Cellules fusionnées verticalement (xlrd). Coordonnées 0-based."""
    continuation: set[tuple[int, int]] = set()
    if not hasattr(sheet, "merged_cells"):
        return continuation
    for rlo, rhi, clo, chi in sheet.merged_cells:
        if rhi - rlo <= 1:
            continue
        for r in range(rlo + 1, rhi):
            for c in range(clo, chi):
                continuation.add((r, c))
    return continuation


def _xls_sheet_to_rows(sheet) -> SheetRows:
    """Feuille xlrd → liste de lignes avec marqueurs de fusion."""
    continuation = _xls_merged_continuation_cells(sheet)
    nrows = sheet.nrows
    ncols = sheet.ncols
    if nrows == 0 or ncols == 0:
        return []

    rows: SheetRows = []
    for r in range(nrows):
        row_vals = []
        for c in range(ncols):
            if (r, c) in continuation:
                row_vals.append("#")
            else:
                row_vals.append(sheet.cell_value(r, c))
        rows.append(row_vals)
    return rows


def _xls_sheet_to_markdown(sheet) -> str:
    """Feuille xlrd en markdown."""
    rows = _xls_sheet_to_rows(sheet)
    return _rows_to_markdown(rows)


# Namespaces ODF (OASIS) pour parser content.xml
_ODF_NS = {
    "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
    "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
    "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
}
_TABLE_TABLE = "{%s}table" % _ODF_NS["table"]
_TABLE_ROW = "{%s}table-row" % _ODF_NS["table"]
_TABLE_CELL = "{%s}table-cell" % _ODF_NS["table"]
_TABLE_COVERED = "{%s}covered-table-cell" % _ODF_NS["table"]
_OFFICE_VALUE = "{%s}value" % _ODF_NS["office"]
_TEXT_P = "{%s}p" % _ODF_NS["text"]


def _ods_cell_text_elt(cell_elt) -> str:
    """Extrait le texte d'une cellule depuis un élément XML (table:table-cell)."""
    val = cell_elt.get(_OFFICE_VALUE)
    if val is not None and str(val).strip():
        return str(val).strip()
    for p in cell_elt.findall(f".//{_TEXT_P}"):
        text = "".join(p.itertext()).strip()
        if text:
            return text
    return ""


def _ods_parse_sheet(table_elt) -> tuple[str, SheetRows]:
    """Parse un élément table:table ; retourne (nom, lignes)."""
    name = table_elt.get("{%s}name" % _ODF_NS["table"], "") or ""
    rows: SheetRows = []
    for row_elt in table_elt.findall(f".//{_TABLE_ROW}"):
        line = []
        for child in row_elt:
            if child.tag == _TABLE_CELL:
                line.append(_ods_cell_text_elt(child))
            elif child.tag == _TABLE_COVERED:
                line.append("#")
        if line:
            rows.append(line)
    return name, rows


def _ods_parse_content(content_xml: bytes) -> list[tuple[str, SheetRows]]:
    """Parse content.xml d'un ODS ; retourne [(nom_feuille, rows), ...]."""
    root = ET.fromstring(content_xml)
    result = []
    for table_elt in root.iter(_TABLE_TABLE):
        result.append(_ods_parse_sheet(table_elt))
    return result


def extract_text_from_xlsx(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier XLSX à partir de son contenu binaire.

    Args:
        file_content: Contenu du fichier .xlsx
        file_path: Chemin ou nom du fichier (pour les logs)
        sep: Séparateur entre les feuilles

    Returns:
        (texte markdown, False) — pas d'OCR pour Excel
    """
    wb = load_workbook(io.BytesIO(file_content), read_only=False, data_only=True)
    sheet_names = wb.sheetnames
    parts = [MERGE_LEGEND]
    for name in sheet_names:
        ws = wb[name]
        md = _xlsx_sheet_to_markdown(ws)
        if md:
            parts.append(f"## {name}\n\n{md}")
    wb.close()
    return sep.join(parts), False


def extract_text_from_ods(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier ODS à partir de son contenu binaire.
    Utilise uniquement la stdlib : zipfile + xml.etree (pas de pandas ni odfpy).

    Returns:
        (texte markdown, False)
    """
    with zipfile.ZipFile(io.BytesIO(file_content), "r") as z:
        content_xml = z.read("content.xml")
    sheets = _ods_parse_content(content_xml)
    parts = [MERGE_LEGEND]
    for name, rows in sheets:
        md = _rows_to_markdown(rows)
        if md:
            parts.append(f"## {name}\n\n{md}")
    return sep.join(parts), False


def extract_text_from_xls(file_content: bytes, file_path: str = "", sep: str = "\n\n") -> tuple[str, bool]:
    """
    Extrait le texte (markdown) d'un fichier XLS à partir de son contenu binaire.

    Returns:
        (texte markdown, False)
    """
    try:
        wb = xlrd.open_workbook(file_contents=file_content, formatting_info=True)
    except (xlrd.XLRDError, NotImplementedError):
        wb = xlrd.open_workbook(file_contents=file_content, formatting_info=False)

    sheet_names = wb.sheet_names()
    parts = [MERGE_LEGEND]
    for name in sheet_names:
        sheet = wb.sheet_by_name(name)
        md = _xls_sheet_to_markdown(sheet)
        if md:
            parts.append(f"## {name}\n\n{md}")
    return sep.join(parts), False
